# app.py  –  Streamlit front-end for the air-cargo cleaning workflow
import streamlit as st

# MUST be the very first Streamlit command
st.set_page_config(page_title="ClearCargo360", layout="wide")

# Now import everything else
import io
import base64
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import yaml
import json
import re

# ------------------------------------------------------------------ #
# Utility helpers (FIXED VERSION with European Air Transport filter)
# ------------------------------------------------------------------ #
DATE_FMT = "%d-%b-%Y"

# FIXED: Better regex to capture airline prefixes from call signs
CALL_SIGN_RE = re.compile(r"^([A-Za-z]+[0-9]*|[0-9]+[A-Za-z]*)(?=[-0-9])?")

# FIXED: Updated regex to capture common airline prefix patterns
CALL_RE = re.compile(r"^([A-Za-z]{1,3}|[0-9]{1,2}[A-Za-z]{1,2}|[A-Za-z]{1,2}[0-9]{1,2})")


def extract_prefix(cs: str) -> str | None:
    """Extract airline prefix from call sign with improved pattern matching"""
    if not isinstance(cs, str):
        return None
    cs = cs.strip().upper()

    # Handle common patterns for airline prefixes:
    # UR-900, UR0900 -> UR (keep original call sign)
    # P47579 -> P4, QY081 -> QY, ME571 -> ME
    # AFR123 -> AFR, BCS456 -> BCS

    # First, try to find letter sequences at the start
    # Look for 2-3 letters followed by numbers, hyphens, or end of string
    match = re.match(r'^([A-Za-z]{2,3})(?=[-0-9]|$)', cs)
    if match:
        return match.group(1).upper()

    # If no match, try 2 characters (handles cases like P4, 8V)
    if len(cs) >= 2:
        prefix = cs[:2]
        # Accept if it's letters, or letter+digit, or digit+letter
        if (prefix.isalpha() or
                (prefix[0].isalpha() and prefix[1].isdigit()) or
                (prefix[0].isdigit() and prefix[1].isalpha())):
            return prefix.upper()

    return None


def handle_numeric_callsigns(df: pd.DataFrame) -> pd.DataFrame:
    """Handle call signs that start with numbers by prepending OperatorAccountingCode"""
    call_col = "CallSign_FlightNo"
    acc_code_col = "OperatorAccountingCode"

    # Check if OperatorAccountingCode column exists
    if acc_code_col not in df.columns:
        return df

    # Find rows where call sign starts with a number (like 1407A, 123B, etc.)
    starts_with_number_mask = df[call_col].str.match(r'^[0-9]', na=False)

    # For call signs starting with numbers, prepend the accounting code
    df.loc[starts_with_number_mask, call_col] = (
            df.loc[starts_with_number_mask, acc_code_col].astype(str) +
            df.loc[starts_with_number_mask, call_col].astype(str)
    )

    return df


def filter_european_air_transport(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filter European Air Transport records to only include AWBIssuingAirline >= 155
    Returns: (filtered_df, excluded_df)
    """
    awb_air_col = "AWBIssuingAirline"
    op_col = "OperatorName"

    # Identify European Air Transport records (case-insensitive)
    european_air_mask = df[op_col].str.upper().str.contains(
        'EUROPEAN AIR TRANSPORT', na=False, regex=False
    )

    if not european_air_mask.any():
        # No European Air Transport records found, return original dataframe
        return df, pd.DataFrame()

    # Convert AWBIssuingAirline to numeric for comparison
    df_copy = df.copy()
    df_copy[awb_air_col + '_numeric'] = pd.to_numeric(df_copy[awb_air_col], errors='coerce')

    # Create filter for European Air Transport records with AWB < 155
    european_exclude_mask = (
            european_air_mask &
            (df_copy[awb_air_col + '_numeric'] < 155)
    )

    # Separate excluded records for reporting
    excluded_df = df[european_exclude_mask].copy()

    # Keep all records except excluded European Air Transport ones
    filtered_df = df[~european_exclude_mask].copy()

    return filtered_df, excluded_df


def normalise_date(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    fmt = parsed.dt.strftime(DATE_FMT)
    # Capitalize only first letter of month: JUN -> Jun, AUG -> Aug
    fmt = fmt.str.replace(r'-([A-Z]{3})-', lambda m: f'-{m.group(1).capitalize()}-', regex=True)
    return fmt.where(parsed.notna(), series)


def load_mapping_and_alias(uploaded_file) -> tuple[dict, dict]:
    """
    Accept a JSON/YAML file that EITHER:
      • contains {'mapping': {...}, 'alias': {...}}
      • OR is a flat {prefix: airline} object (then alias = {})
    Returns (modern_map, alias_map)
    """
    suffix = Path(uploaded_file.name).suffix.lower()
    data = yaml.safe_load(uploaded_file) if suffix in {".yml", ".yaml"} \
        else json.loads(uploaded_file.read().decode("utf-8"))
    uploaded_file.seek(0)  # reset for any future reads

    if "mapping" in data:
        return data["mapping"], data.get("alias", {})
    return data, {}  # simple flat file


def clean_and_split(df: pd.DataFrame, mapping: dict, alias: dict) -> tuple[
    dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    # --- normalise headers --------------------------------------------------
    df.columns = [c.strip() for c in df.columns]

    call_col, op_col = "CallSign_FlightNo", "OperatorName"
    awb_air, awb_ser, wt = "AWBIssuingAirline", "AWBSerialNumber", "CargoWeight"

    # --- handle numeric-only call signs first -------------------------------
    df = handle_numeric_callsigns(df)

    # --- derive prefix & modern prefix --------------------------------------
    df["__LEG"] = df[call_col].apply(extract_prefix)
    df["__MOD"] = df["__LEG"].map(alias).fillna(df["__LEG"])  # map legacy→modern using alias

    # --- rename operator ONLY (do NOT change call sign) --------------------
    df[op_col] = (
        df["__MOD"].map(mapping)  # look up airline by modern prefix
        .where(lambda x: x.notna(), df[op_col])  # fallback to existing OperatorName
        .str.upper()
    )

    # --- standardize call signs using modern prefixes ----------------------
    # Replace original prefixes with their modern equivalents
    # E.g., AF0132 -> AFR0132, QR1407A -> QTR1407A, UR900 -> UGD900
    def standardize_callsign(cs: str, old_prefix: str, modern_prefix: str) -> str:
        if pd.isna(old_prefix) or pd.isna(modern_prefix) or old_prefix == modern_prefix:
            return cs
        if cs.startswith(old_prefix):
            return cs.replace(old_prefix, modern_prefix, 1)
        return cs

    df[call_col] = [
        standardize_callsign(cs, old, new)
        for cs, old, new in zip(df[call_col], df["__LEG"], df["__MOD"])
    ]

    # --- normalise any *DATE* columns ---------------------------------------
    for col in (c for c in df.columns if "DATE" in c.upper()):
        df[col] = normalise_date(df[col])

    # --- additional operator name cleaning -----------------------------------
    # Handle common variations that might not be caught by prefix mapping
    operator_fixes = {
        'QATAR': 'QATAR AIRWAYS',
        'EMIRATES AIRLINES': 'EMIRATES AIRLINE',
        'LUFTHANSA GERMAN AIRLINES': 'DEUTSCHE LUFTHANSA AG',
        'DELTA AIRLINE': 'DELTA AIR LINES INC',
        'KLM ROYAL': 'KLM ROYAL DUTCH AIRLINES',
        'EGYPT': 'EGYPTAIR',
        'DHL INTERNATIONAL NIGERIA LIMITED': 'EUROPEAN AIR TRANSPORT',
        'AIRPEACE': 'AIR PEACE LIMITED',  # Standardize Air Peace variations
        'AIR PEACE': 'AIR PEACE LIMITED'
    }

    # Apply exact matches only
    for old_name, new_name in operator_fixes.items():
        df[op_col] = df[op_col].str.replace(old_name, new_name, regex=False)

    # --- APPLY EUROPEAN AIR TRANSPORT FILTER --------------------------------
    df_filtered, excluded_european = filter_european_air_transport(df)

    # --- capture unmapped prefixes ------------------------------------------
    # FIXED: Only include rows where the modern prefix is NOT in the mapping
    unmapped = df_filtered[
        df_filtered["__MOD"].notna() &
        ~df_filtered["__MOD"].isin(mapping) &
        df_filtered["__LEG"].notna()  # Only include rows with valid extracted prefixes
        ].copy()

    # --- deduplicate per airline --------------------------------------------
    df_filtered["__WB"] = df_filtered[awb_air].astype(str) + "-" + df_filtered[awb_ser].astype(str)
    cleaned: dict[str, pd.DataFrame] = {}

    # Only process rows that have valid mappings
    mapped_df = df_filtered[df_filtered["__MOD"].isin(mapping)].copy()

    for name, grp in mapped_df.groupby(op_col, sort=True):
        dedup = grp[~grp.duplicated(["__WB", wt])].copy()
        dedup.drop(columns=["__LEG", "__MOD", "__WB"], inplace=True)
        cleaned[name] = dedup.reset_index(drop=True)

    return cleaned, unmapped, excluded_european


def build_multi_sheet_excel(data: dict[str, pd.DataFrame], unmapped: pd.DataFrame,
                            excluded_european: pd.DataFrame = None) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as w:
        # Create combined sheet first with all data
        combined_data = []
        for airline_name, frame in data.items():
            combined_data.append(frame)

        if combined_data:
            combined_df = pd.concat(combined_data, ignore_index=True)
            combined_df.to_excel(w, sheet_name="Combined", index=False)

        # Create individual airline sheets
        for name, frame in data.items():
            frame.to_excel(w, sheet_name=name[:31], index=False)

        # Add unmapped sheet if there are unmapped rows
        if not unmapped.empty:
            unmapped.to_excel(w, sheet_name="Unmapped", index=False)

        # Add excluded European Air Transport sheet if there are excluded rows
        if excluded_european is not None and not excluded_european.empty:
            excluded_european.to_excel(w, sheet_name="Excluded_European_Air", index=False)

    return buffer.getvalue()


def fill_iata(template_bytes: bytes, cleaned: dict[str, pd.DataFrame],
              sheet: str = "CARGO SALES") -> bytes:
    buffer = io.BytesIO(template_bytes)
    wb = load_workbook(buffer, keep_vba=True)
    ws = wb[sheet]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for frame in cleaned.values():
        for row in frame.itertuples(index=False):
            ws.append(list(row))
    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()


def create_duplicate_report(dups1: pd.DataFrame, dups2: pd.DataFrame,
                            file1_name: str, file2_name: str, comparison_method: str) -> bytes:
    """Create a detailed Excel report of duplicate records between two files"""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as w:
        # Summary sheet
        summary_data = {
            'Metric': [
                'File 1 Name', 'File 2 Name', 'Comparison Method',
                'Duplicates in File 1', 'Duplicates in File 2', 'Analysis Date'
            ],
            'Value': [
                file1_name, file2_name, comparison_method,
                len(dups1), len(dups2), pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(w, sheet_name="Summary", index=False)

        # Duplicates from file 1
        if not dups1.empty:
            dups1.to_excel(w, sheet_name=f"Duplicates_from_{file1_name[:15]}", index=False)

        # Duplicates from file 2
        if not dups2.empty:
            dups2.to_excel(w, sheet_name=f"Duplicates_from_{file2_name[:15]}", index=False)

    return buffer.getvalue()


def b64_download(data: bytes, filename: str, label: str):
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" ' \
           f'download="{filename}">{label}</a>'
    st.markdown(href, unsafe_allow_html=True)


# FIXED: Improved header detection function
def detect_has_headers(df_sample: pd.DataFrame) -> bool:
    """
    Improved logic to detect if a dataframe has proper headers
    Returns True if headers are detected, False otherwise
    """
    if df_sample.empty:
        return True  # Default to assuming headers

    # Check 1: Are column names meaningful (contain expected keywords)?
    header_keywords = ['awb', 'operator', 'cargo', 'flight', 'date', 'weight', 'airline', 'serial']
    meaningful_headers = 0
    for col in df_sample.columns:
        col_str = str(col).lower()
        if any(keyword in col_str for keyword in header_keywords):
            meaningful_headers += 1

    if meaningful_headers >= 2:  # At least 2 meaningful headers
        return True

    # Check 2: Are columns just numbers (0, 1, 2...)?
    if all(isinstance(col, (int, float)) for col in df_sample.columns):
        return False

    # Check 3: Are columns like 'Unnamed: 0', 'Unnamed: 1'?
    if all('unnamed' in str(col).lower() for col in df_sample.columns):
        return False

    # Check 4: Does first row contain typical airline data?
    if len(df_sample) > 0:
        first_row = df_sample.iloc[0]
        airline_patterns = [
            'AIR FRANCE', 'EMIRATES', 'QATAR', 'LUFTHANSA', 'DELTA',
            'KLM', 'TURKISH', 'ETHIOPIAN', 'BRITISH', 'EUROPEAN AIR',
            'AIR PEACE', 'VIRGIN', 'SOUTHWEST', 'AMERICAN'
        ]

        # Check if any cell in first row contains airline names
        for val in first_row:
            val_str = str(val).upper()
            if any(airline in val_str for airline in airline_patterns):
                return False  # First row contains data, not headers

        # Check if first row has numeric patterns typical of AWB data
        numeric_count = sum(1 for val in first_row if str(val).replace('.', '').replace('-', '').isdigit())
        if numeric_count >= 3:  # If 3+ numeric values, likely data row
            return False

    # Default to True if unclear
    return True


def read_sheet_with_proper_headers(file, sheet_name):
    """Read Excel sheet and intelligently handle missing headers"""
    # Standard column names for air cargo data
    expected_columns = [
        'OperatorName', 'OperatorEFLegacyCode', 'OperatorAccountingCode',
        'FlightDate', 'MonthCheckDigitManifest', 'CallSign_FlightNo',
        'FromAirportCode', 'ToAirportCode', 'AWBDate', 'MonthCheckDigitAWB',
        'AWBIssuingAirline', 'AWBSerialNumber', 'AWBCheckDigit', 'CargoWeight',
        'RateKGM', 'FeesCargoTotalAmount', 'CurrencyCode', 'SyncedAt'
    ]

    # First, read a sample to inspect headers
    df_sample = pd.read_excel(file, sheet_name=sheet_name, nrows=3)
    has_proper_headers = detect_has_headers(df_sample)

    # Now read the file properly
    if has_proper_headers:
        # File has headers, read normally
        df = pd.read_excel(file, sheet_name=sheet_name)
    else:
        # File doesn't have headers, read without header row
        df = pd.read_excel(file, sheet_name=sheet_name, header=None)

        # Assign standard column names
        num_cols = len(df.columns)
        if num_cols >= len(expected_columns):
            df.columns = expected_columns + [f'Extra_{i}' for i in range(num_cols - len(expected_columns))]
        else:
            df.columns = expected_columns[:num_cols]

    return df, has_proper_headers


# FIXED: Enhanced column finder with better fuzzy matching
def find_best_match(df_columns, target_name, position_fallback=None):
    """Find best matching column name with position fallback and fuzzy matching"""

    # Define comprehensive column name variations
    column_variations = {
        "AWBIssuingAirline": [
            "awbissuingairline", "awb_issuing_airline", "awb issuing airline",
            "issuing airline", "awb airline", "awbairline", "awb_airline",
            "airway bill issuing airline", "awb issuing", "issuing_airline"
        ],
        "AWBSerialNumber": [
            "awbserialnumber", "awb_serial_number", "awb serial number",
            "serial number", "awb serial", "awbserial", "awb_serial",
            "airway bill serial", "awb_number", "awbnumber", "awb number"
        ],
        "CargoWeight": [
            "cargoweight", "cargo_weight", "cargo weight", "weight",
            "cargo wt", "wt", "cargo_wt", "weight_kg", "weightkg",
            "cargo_weight_kg", "freight weight", "shipment weight"
        ],
        "OperatorName": [
            "operatorname", "operator_name", "operator name", "airline",
            "airline name", "carrier", "operator", "airline_name",
            "carrier_name", "aviation operator"
        ]
    }

    # Step 1: Exact match (case insensitive)
    for col in df_columns:
        if str(col).lower().strip() == target_name.lower():
            return col

    # Step 2: Try predefined variations
    variations = column_variations.get(target_name, [target_name.lower()])
    for variation in variations:
        for col in df_columns:
            col_clean = str(col).lower().strip().replace(' ', '').replace('_', '')
            variation_clean = variation.replace(' ', '').replace('_', '')
            if variation_clean in col_clean or col_clean in variation_clean:
                return col

    # Step 3: Position fallback if provided
    if position_fallback is not None and 0 <= position_fallback < len(df_columns):
        return df_columns[position_fallback]

    return None


# FIXED: Normalize data for comparison
def normalize_awb_data(df, awb_airline_col, awb_serial_col, weight_col=None):
    """Normalize AWB data for consistent comparison"""
    df_norm = df.copy()

    # Normalize AWB Airline - convert to string, strip whitespace, handle NaN
    df_norm[awb_airline_col] = (df_norm[awb_airline_col]
                                .fillna('')
                                .astype(str)
                                .str.strip()
                                .str.replace(r'\.0$', '', regex=True))  # Remove .0 from floats

    # Normalize AWB Serial - more robust handling
    df_norm[awb_serial_col] = (df_norm[awb_serial_col]
                               .fillna('')
                               .astype(str)
                               .str.strip()
                               .str.replace(r'\.0$', '', regex=True)  # Remove .0 from floats
                               .str.lstrip('0')  # Remove leading zeros
                               .replace('', '0'))  # Replace empty strings with '0'

    # Normalize weight if provided - more robust handling
    if weight_col and weight_col in df_norm.columns:
        df_norm[weight_col] = pd.to_numeric(df_norm[weight_col], errors='coerce').fillna(0).round(2)

    return df_norm


def create_comparison_key(airline, serial, weight=None):
    """Create a standardized comparison key"""
    # Clean and normalize components
    airline_clean = str(airline).strip().replace('.0', '')
    serial_clean = str(serial).strip().replace('.0', '').lstrip('0') or '0'

    if weight is not None:
        weight_clean = str(float(weight)).replace('.0', '') if pd.notna(weight) else '0'
        return f"{airline_clean}-{serial_clean}-{weight_clean}"
    else:
        return f"{airline_clean}-{serial_clean}"


# -----------------------  Streamlit UI  ------------------------------------ #
st.title("📦 ClearCargo 360")
st.markdown("#### *From manifest to money — reconciling air-cargo data for accurate billing.*")

# Create tabs for different functions
tab1, tab2 = st.tabs(["🔧 Reconcile Data", "🔍 Billing Conflict Finder"])

with tab1:
    st.markdown("### Clean and Process Air Cargo Data")

    raw_file = st.file_uploader("Raw Data Excel", type="xlsx", key="clean_raw")
    map_file = st.file_uploader("Call-sign mapping (YAML or JSON)", type=("yml", "yaml", "json"), key="clean_map")
    iata_file = st.file_uploader("IATA .xlsm template (single sheet)", type="xlsm", key="clean_iata")

    if st.button("Process", disabled=not (raw_file and map_file and iata_file)):
        try:
            # FIXED: Properly unpack both mapping and alias
            mapping, alias = load_mapping_and_alias(map_file)
            raw_df = pd.read_excel(raw_file, sheet_name="MainSheet")

            # FIXED: Pass both mapping and alias to the function and handle excluded European records
            cleaned, unmapped, excluded_european = clean_and_split(raw_df, mapping, alias)

            # Build outputs
            sheet_book = build_multi_sheet_excel(cleaned, unmapped, excluded_european)
            iata_ready = fill_iata(iata_file.read(), cleaned)

            st.success("Done! Download your files below.")

            # Display some debug info
            st.write("**Processing Summary:**")
            st.write(f"- Total airlines processed: {len(cleaned)}")
            st.write(f"- Airlines: {', '.join(cleaned.keys())}")

            # Calculate total rows in combined data
            total_rows = sum(len(df) for df in cleaned.values())
            st.write(f"- Total rows in combined sheet: {total_rows}")

            # Show European Air Transport filter results
            if not excluded_european.empty:
                st.warning(
                    f"⚠️ Excluded {len(excluded_european)} European Air Transport records with AWBIssuingAirline < 155")
                st.write("**Sample Excluded European Air Transport Records:**")
                display_cols = ["AWBIssuingAirline", "OperatorName", "CallSign_FlightNo"]
                if all(col in excluded_european.columns for col in display_cols):
                    st.dataframe(excluded_european[display_cols].head(5))
            else:
                st.info(
                    "✅ No European Air Transport records excluded (all AWB >= 155 or no European Air Transport records found)")

            # Debug: Check specific call signs
            call_col = "CallSign_FlightNo"  # Define it here for debugging
            if 'UR' in [extract_prefix(cs) for cs in raw_df[call_col].dropna()[:100]]:
                ur_rows = raw_df[raw_df[call_col].str.contains('UR', na=False)]
                if not ur_rows.empty:
                    st.write("**🔍 UR Call Signs Found:**")
                    ur_sample = ur_rows[[call_col, "OperatorName"]].head(5)
                    st.dataframe(ur_sample)

            if not unmapped.empty:
                unique_unmapped = unmapped['__LEG'].unique()
                st.warning(f"{len(unmapped)} rows have unmapped prefixes: {', '.join(unique_unmapped)}")

                # Show sample unmapped for debugging
                st.write("**Sample Unmapped Rows:**")
                unmapped_sample = unmapped[["CallSign_FlightNo", "OperatorName", '__LEG', '__MOD']].head(5)
                st.dataframe(unmapped_sample)
            else:
                st.success("✅ No unmapped data found!")

            b64_download(sheet_book, "airline_sheets.xlsx", "⬇️ Airline Sheets (XLSX)")
            b64_download(iata_ready, "iata_ready.xlsm", "⬇️ IATA Template (XLSM)")

        except Exception as e:
            st.error(f"Processing failed: {e}")
            st.exception(e)  # Show full traceback for debugging

with tab2:
    st.markdown("### Check for Matches Between Two Files ")
    st.markdown("*Compare two Excel files to find matching records based on AWB (waybill) numbers and weights.*")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**📅 First Period (e.g., June 1-7)**")
        file1 = st.file_uploader("First Excel file", type="xlsx", key="dup_file1")
        if file1:
            sheet1 = st.selectbox("Select sheet from first file",
                                  options=pd.ExcelFile(file1).sheet_names,
                                  key="sheet1")

    with col2:
        st.markdown("**📅 Second Period (e.g., June 8-17)**")
        file2 = st.file_uploader("Second Excel file", type="xlsx", key="dup_file2")
        if file2:
            sheet2 = st.selectbox("Select sheet from second file",
                                  options=pd.ExcelFile(file2).sheet_names,
                                  key="sheet2")

    # Duplicate checking options
    st.markdown("**🔧 Duplicate Detection Settings**")
    check_weight = st.checkbox("Include weight in duplicate check", value=True,
                               help="If checked, rows must have same AWB AND weight to be considered duplicates")

    # Header configuration options
    st.markdown("**📋 Header Configuration**")
    col1, col2 = st.columns(2)
    with col1:
        force_no_header1 = st.checkbox("File 1 has no headers", key="no_header1",
                                       help="Check if first row contains data, not column names")
    with col2:
        force_no_header2 = st.checkbox("File 2 has no headers", key="no_header2",
                                       help="Check if first row contains data, not column names")

    # Preview button
    if file1 and file2:
        if st.button("Preview Data Structure", type="secondary"):
            st.markdown("### 📊 Data Preview")

            # Read files with header detection
            try:
                if force_no_header1:
                    df1_preview = pd.read_excel(file1, sheet_name=sheet1, header=None, nrows=3)
                    expected_columns = [
                        'OperatorName', 'OperatorEFLegacyCode', 'OperatorAccountingCode',
                        'FlightDate', 'MonthCheckDigitManifest', 'CallSign_FlightNo',
                        'FromAirportCode', 'ToAirportCode', 'AWBDate', 'MonthCheckDigitAWB',
                        'AWBIssuingAirline', 'AWBSerialNumber', 'AWBCheckDigit', 'CargoWeight',
                        'RateKGM', 'FeesCargoTotalAmount', 'CurrencyCode', 'SyncedAt'
                    ]
                    df1_preview.columns = expected_columns[:len(df1_preview.columns)]
                    has_headers1 = False
                else:
                    df1_preview, has_headers1 = read_sheet_with_proper_headers(file1, sheet1)
                    df1_preview = df1_preview.head(3)

                if force_no_header2:
                    df2_preview = pd.read_excel(file2, sheet_name=sheet2, header=None, nrows=3)
                    expected_columns = [
                        'OperatorName', 'OperatorEFLegacyCode', 'OperatorAccountingCode',
                        'FlightDate', 'MonthCheckDigitManifest', 'CallSign_FlightNo',
                        'FromAirportCode', 'ToAirportCode', 'AWBDate', 'MonthCheckDigitAWB',
                        'AWBIssuingAirline', 'AWBSerialNumber', 'AWBCheckDigit', 'CargoWeight',
                        'RateKGM', 'FeesCargoTotalAmount', 'CurrencyCode', 'SyncedAt'
                    ]
                    df2_preview.columns = expected_columns[:len(df2_preview.columns)]
                    has_headers2 = False
                else:
                    df2_preview, has_headers2 = read_sheet_with_proper_headers(file2, sheet2)
                    df2_preview = df2_preview.head(3)

                # Show previews with column indices and header detection results
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown(
                        f"**File 1 Structure** ({'Headers Detected' if has_headers1 else 'No Headers Detected'}):")
                    preview_df1 = df1_preview.copy()
                    preview_df1.columns = [f"[{i}] {col}" for i, col in enumerate(preview_df1.columns)]
                    st.dataframe(preview_df1)

                with col2:
                    st.markdown(
                        f"**File 2 Structure** ({'Headers Detected' if has_headers2 else 'No Headers Detected'}):")
                    preview_df2 = df2_preview.copy()
                    preview_df2.columns = [f"[{i}] {col}" for i, col in enumerate(preview_df2.columns)]
                    st.dataframe(preview_df2)

            except Exception as e:
                st.error(f"Error previewing files: {e}")

    if st.button("🔍 Find Duplicates", disabled=not (file1 and file2)):
        try:
            # Read files with improved header detection
            if force_no_header1:
                df1 = pd.read_excel(file1, sheet_name=sheet1, header=None)
                expected_columns = [
                    'OperatorName', 'OperatorEFLegacyCode', 'OperatorAccountingCode',
                    'FlightDate', 'MonthCheckDigitManifest', 'CallSign_FlightNo',
                    'FromAirportCode', 'ToAirportCode', 'AWBDate', 'MonthCheckDigitAWB',
                    'AWBIssuingAirline', 'AWBSerialNumber', 'AWBCheckDigit', 'CargoWeight',
                    'RateKGM', 'FeesCargoTotalAmount', 'CurrencyCode', 'SyncedAt'
                ]
                num_cols = len(df1.columns)
                if num_cols >= len(expected_columns):
                    df1.columns = expected_columns + [f'Extra_{i}' for i in range(num_cols - len(expected_columns))]
                else:
                    df1.columns = expected_columns[:num_cols]
                has_headers1 = False
            else:
                df1, has_headers1 = read_sheet_with_proper_headers(file1, sheet1)

            if force_no_header2:
                df2 = pd.read_excel(file2, sheet_name=sheet2, header=None)
                expected_columns = [
                    'OperatorName', 'OperatorEFLegacyCode', 'OperatorAccountingCode',
                    'FlightDate', 'MonthCheckDigitManifest', 'CallSign_FlightNo',
                    'FromAirportCode', 'ToAirportCode', 'AWBDate', 'MonthCheckDigitAWB',
                    'AWBIssuingAirline', 'AWBSerialNumber', 'AWBCheckDigit', 'CargoWeight',
                    'RateKGM', 'FeesCargoTotalAmount', 'CurrencyCode', 'SyncedAt'
                ]
                num_cols = len(df2.columns)
                if num_cols >= len(expected_columns):
                    df2.columns = expected_columns + [f'Extra_{i}' for i in range(num_cols - len(expected_columns))]
                else:
                    df2.columns = expected_columns[:num_cols]
                has_headers2 = False
            else:
                df2, has_headers2 = read_sheet_with_proper_headers(file2, sheet2)

            # Display header status
            col1, col2 = st.columns(2)
            with col1:
                if has_headers1 or force_no_header1:
                    st.success(f"✓ File 1: {'Forced no headers' if force_no_header1 else 'Headers detected'}")
                else:
                    st.warning(f"⚠️ File 1: No headers detected - applied standard columns")

            with col2:
                if has_headers2 or force_no_header2:
                    st.success(f"✓ File 2: {'Forced no headers' if force_no_header2 else 'Headers detected'}")
                else:
                    st.warning(f"⚠️ File 2: No headers detected - applied standard columns")

            # Debug: Show first few rows of actual data
            # st.markdown("### 🔍 Data Sample Check")
            # col1, col2 = st.columns(2)
            # with col1:
            #     st.markdown("**File 1 Sample:**")
            #     sample_cols = ['AWBIssuingAirline', 'AWBSerialNumber',
            #                    'CargoWeight'] if 'AWBIssuingAirline' in df1.columns else df1.columns[:5]
            #     st.dataframe(df1[sample_cols].head(3))
            # with col2:
            #     st.markdown("**File 2 Sample:**")
            #     sample_cols = ['AWBIssuingAirline', 'AWBSerialNumber',
            #                    'CargoWeight'] if 'AWBIssuingAirline' in df2.columns else df2.columns[:5]
            #     st.dataframe(df2[sample_cols].head(3))

            # Normalize column names
            df1.columns = [str(c).strip() if isinstance(c, str) else str(c) for c in df1.columns]
            df2.columns = [str(c).strip() if isinstance(c, str) else str(c) for c in df2.columns]

            # Enhanced column mapping with position fallback
            position_map = {
                "AWBIssuingAirline": 10,
                "AWBSerialNumber": 11,
                "CargoWeight": 13
            }

            # Map actual column names to expected ones
            required_cols = ["AWBIssuingAirline", "AWBSerialNumber"]
            if check_weight:
                required_cols.append("CargoWeight")

            column_mapping_1 = {}
            column_mapping_2 = {}
            missing_cols_1 = []
            missing_cols_2 = []

            for req_col in required_cols:
                # File 1
                pos_fallback = position_map.get(req_col)
                actual_col_1 = find_best_match(df1.columns, req_col, pos_fallback)
                if actual_col_1:
                    column_mapping_1[req_col] = actual_col_1
                else:
                    missing_cols_1.append(req_col)

                # File 2
                actual_col_2 = find_best_match(df2.columns, req_col, pos_fallback)
                if actual_col_2:
                    column_mapping_2[req_col] = actual_col_2
                else:
                    missing_cols_2.append(req_col)

            # # Show column mapping results
            # st.markdown("### 📋 Column Mapping Results")
            # col1, col2 = st.columns(2)
            # with col1:
            #     st.markdown("**File 1 Mapping:**")
            #     for req, actual in column_mapping_1.items():
            #         st.success(f"✓ {req} → {actual}")
            #     if missing_cols_1:
            #         for missing in missing_cols_1:
            #             st.error(f"✗ {missing} → NOT FOUND")
            #
            # with col2:
            #     st.markdown("**File 2 Mapping:**")
            #     for req, actual in column_mapping_2.items():
            #         st.success(f"✓ {req} → {actual}")
            #     if missing_cols_2:
            #         for missing in missing_cols_2:
            #             st.error(f"✗ {missing} → NOT FOUND")

            # If still missing columns after position fallback, show error
            if missing_cols_1 or missing_cols_2:
                st.error("❌ Unable to find all required columns")
                st.warning("💡 Try checking the 'has no headers' option if columns aren't being detected correctly")

                # Show available columns for debugging
                st.markdown("**Available Columns for Reference:**")
                col1, col2 = st.columns(2)
                with col1:
                    st.write("File 1:", list(df1.columns))
                with col2:
                    st.write("File 2:", list(df2.columns))
            else:
                # Use mapped column names for comparison
                awb_air_1 = column_mapping_1["AWBIssuingAirline"]
                awb_ser_1 = column_mapping_1["AWBSerialNumber"]
                awb_air_2 = column_mapping_2["AWBIssuingAirline"]
                awb_ser_2 = column_mapping_2["AWBSerialNumber"]

                # Normalize data before comparison
                df1_norm = normalize_awb_data(df1, awb_air_1, awb_ser_1,
                                              column_mapping_1.get("CargoWeight") if check_weight else None)
                df2_norm = normalize_awb_data(df2, awb_air_2, awb_ser_2,
                                              column_mapping_2.get("CargoWeight") if check_weight else None)

                # Create comparison keys using the improved function
                if check_weight:
                    wt_1 = column_mapping_1["CargoWeight"]
                    wt_2 = column_mapping_2["CargoWeight"]

                    # Create keys using the create_comparison_key function
                    df1_keys = []
                    df2_keys = []

                    for _, row in df1_norm.iterrows():
                        key = create_comparison_key(row[awb_air_1], row[awb_ser_1], row[wt_1])
                        df1_keys.append(key)

                    for _, row in df2_norm.iterrows():
                        key = create_comparison_key(row[awb_air_2], row[awb_ser_2], row[wt_2])
                        df2_keys.append(key)

                    df1_norm["__KEY"] = df1_keys
                    df2_norm["__KEY"] = df2_keys
                    key_description = "AWB + Weight"
                else:
                    # Create keys using only AWB data
                    df1_keys = []
                    df2_keys = []

                    for _, row in df1_norm.iterrows():
                        key = create_comparison_key(row[awb_air_1], row[awb_ser_1])
                        df1_keys.append(key)

                    for _, row in df2_norm.iterrows():
                        key = create_comparison_key(row[awb_air_2], row[awb_ser_2])
                        df2_keys.append(key)

                    df1_norm["__KEY"] = df1_keys
                    df2_norm["__KEY"] = df2_keys
                    key_description = "AWB Only"

                # Debug: Show sample keys with components
                # st.markdown("### 🔑 Sample Comparison Keys")
                # col1, col2 = st.columns(2)
                # with col1:
                #     st.markdown("**File 1 Keys:**")
                #     display_cols = ["__KEY", awb_air_1, awb_ser_1]
                #     if check_weight:
                #         display_cols.append(wt_1)
                #     sample_keys_1 = df1_norm[display_cols].head(5)
                #     st.dataframe(sample_keys_1)
                #
                # with col2:
                #     st.markdown("**File 2 Keys:**")
                #     display_cols = ["__KEY", awb_air_2, awb_ser_2]
                #     if check_weight:
                #         display_cols.append(wt_2)
                #     sample_keys_2 = df2_norm[display_cols].head(5)
                #     st.dataframe(sample_keys_2)
                #
                # # Find duplicates with detailed logging
                # st.markdown("### 🔄 Finding Matches...")

                # Get unique keys from both files
                keys_1 = set(df1_norm["__KEY"].dropna())
                keys_2 = set(df2_norm["__KEY"].dropna())

                # Find intersection
                matching_keys = keys_1 & keys_2

                # st.write(f"**Unique keys in File 1:** {len(keys_1)}")
                # st.write(f"**Unique keys in File 2:** {len(keys_2)}")
                # st.write(f"**Matching keys found:** {len(matching_keys)}")

                # Keys comparison logic (hidden for cleaner UI)
                # if len(matching_keys) > 0:
                #     st.success("✅ Found matching keys!")
                #     st.write("**Sample matching keys:**")
                #     for key in list(matching_keys)[:5]:
                #         st.write(f"- `{key}`")
                # else:
                #     st.warning("⚠️ No matching keys found")
                #
                #     # Show some keys from each file for comparison
                #     st.write("**Sample keys from File 1:**")
                #     for key in list(keys_1)[:5]:
                #         st.write(f"- `{key}`")
                #
                #     st.write("**Sample keys from File 2:**")
                #     for key in list(keys_2)[:5]:
                #         st.write(f"- `{key}`")

                # Find actual duplicate rows
                duplicates_in_2 = df2[df2_norm["__KEY"].isin(df1_norm["__KEY"])].copy()
                duplicates_in_1 = df1[df1_norm["__KEY"].isin(df2_norm["__KEY"])].copy()

                st.markdown("---")
                st.markdown("### 📊 Final Results")

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("File 1 Total", len(df1))
                with col2:
                    st.metric("File 2 Total", len(df2))
                # with col3:
                #     st.metric("Matching Keys", len(matching_keys))
                with col3:
                    st.metric("Duplicate Records", len(duplicates_in_1))

                st.info(f"**Comparison Method:** {key_description}")

                if len(duplicates_in_1) > 0:
                    st.success(f"🎯 Found {len(duplicates_in_1)} duplicate records!")

                    # Show sample duplicates with comparison keys
                    st.markdown("### 📋 Matching Records Found")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**Duplicates from File 1:**")
                        display_cols_1 = [awb_air_1, awb_ser_1]
                        if check_weight and wt_1:
                            display_cols_1.append(wt_1)
                        if "OperatorName" in duplicates_in_1.columns:
                            display_cols_1.append("OperatorName")

                        # Add the comparison key to see what matched
                        dup_with_keys_1 = duplicates_in_1.copy()
                        dup_with_keys_1["Comparison_Key"] = df1_norm.loc[duplicates_in_1.index, "__KEY"]
                        display_cols_1.append("Comparison_Key")
                        st.dataframe(dup_with_keys_1[display_cols_1].head(10))

                    with col2:
                        st.markdown("**Duplicates from File 2:**")
                        display_cols_2 = [awb_air_2, awb_ser_2]
                        if check_weight and wt_2:
                            display_cols_2.append(wt_2)
                        if "OperatorName" in duplicates_in_2.columns:
                            display_cols_2.append("OperatorName")

                        # Add the comparison key to see what matched
                        dup_with_keys_2 = duplicates_in_2.copy()
                        dup_with_keys_2["Comparison_Key"] = df2_norm.loc[duplicates_in_2.index, "__KEY"]
                        display_cols_2.append("Comparison_Key")
                        st.dataframe(dup_with_keys_2[display_cols_2].head(10))

                    # Create downloadable report
                    duplicate_report = create_duplicate_report(
                        duplicates_in_1, duplicates_in_2,
                        file1.name, file2.name, key_description
                    )

                    b64_download(
                        duplicate_report,
                        f"duplicate_report_{file1.name}_{file2.name}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        "⬇️ Download Full Duplicate Report"
                    )

                else:
                    st.warning("❌ No matching records found between the files!")

                    # Enhanced debugging when no duplicates found
                    st.markdown("### 🔧 Troubleshooting Information")

                    # Show exact key comparisons
                    st.markdown("**🔍 Key Format Analysis:**")
                    col1, col2 = st.columns(2)

                    with col1:
                        st.write("**File 1 - First 10 keys:**")
                        keys_list_1 = list(keys_1)[:10]
                        for i, key in enumerate(keys_list_1):
                            st.write(f"{i + 1}. `{key}`")

                    with col2:
                        st.write("**File 2 - First 10 keys:**")
                        keys_list_2 = list(keys_2)[:10]
                        for i, key in enumerate(keys_list_2):
                            st.write(f"{i + 1}. `{key}`")

                    # Check for similar keys (debugging)
                    st.markdown("**🔍 Similarity Check:**")
                    similar_found = False
                    sample_keys_1 = list(keys_1)[:5]
                    sample_keys_2 = list(keys_2)[:5]

                    for k1 in sample_keys_1:
                        for k2 in sample_keys_2:
                            # Check if keys are similar (same components but different formatting)
                            k1_parts = k1.split('-')
                            k2_parts = k2.split('-')
                            if len(k1_parts) == len(k2_parts):
                                similarity = sum(1 for a, b in zip(k1_parts, k2_parts) if a == b)
                                if similarity >= len(k1_parts) - 1:  # Allow 1 difference
                                    st.warning(
                                        f"Similar keys found: `{k1}` vs `{k2}` (similarity: {similarity}/{len(k1_parts)})")
                                    similar_found = True

                    if not similar_found:
                        st.info("No similar keys detected. The data in both files appears to be completely different.")

                        # Suggest alternative comparison methods
                        st.markdown("**💡 Suggestions:**")
                        st.write("1. Try unchecking 'Include weight' to compare only AWB numbers")
                        st.write("2. Verify you're comparing the correct sheets")
                        st.write("3. Check if data formatting is consistent between files")
                        st.write("4. Ensure both files contain the same type of records")

        except Exception as e:
            st.error(f"Duplicate check failed: {e}")
            st.exception(e)

# Add a footer
st.markdown("---")
st.markdown("##### Built with ❤️ for air cargo reconciliation")